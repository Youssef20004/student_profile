from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from django.conf import settings
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os
from .models import Student
from import_export.admin import ImportExportModelAdmin
from import_export import resources
from import_export.formats.base_formats import XLSX, CSV

admin.site.site_header = "Dashboard"
admin.site.site_title = "Dashboard Admin"
admin.site.index_title = "Dashboard Admin"

class StudentResource(resources.ModelResource):
    class Meta:
        model = Student
        fields = ('seat_number', 'code', 'arabic_name', 'national_id', 'english_name', 'photo')
        export_order = ('seat_number', 'code', 'arabic_name', 'national_id', 'english_name', 'photo')
        import_id_fields = ('seat_number',)  # الحقل اللي بيحدد السجل الفريد أثناء الاستيراد

    def before_import_row(self, row, **kwargs):
        # تحويل القيم الفارغة إلى None لتجنب الأخطاء
        for key in row:
            if row[key] == '':
                row[key] = None

    def after_import_instance(self, instance, new, **kwargs):
        # التحقق من حقل الصورة إذا كانت موجودة
        if instance.photo and not os.path.exists(os.path.join(settings.MEDIA_ROOT, str(instance.photo))):
            instance.photo = None  # إزالة المسار إذا الصورة غير موجودة

@admin.register(Student)
class StudentAdmin(ImportExportModelAdmin):
    resource_class = StudentResource  # ربط الكلاس بالـ Resource
    formats = [XLSX, CSV]  # تحديد صيغ الملفات المسموح بها للاستيراد والتصدير
    skip_import_confirm = True  # تخطي صفحة تأكيد الاستيراد
    list_display = ('arabic_name', 'code', 'seat_number', 'national_id', 'english_name', 'display_photo', 'has_english_name')
    list_filter = ('english_name', 'created')
    search_fields = ('seat_number', 'arabic_name', 'english_name', 'code', 'national_id')
    readonly_fields = ('display_photo', 'created')
    actions = ['export_to_excel']
    fieldsets = (
        ('Basic Information', {
            'fields': ('seat_number', 'code', 'arabic_name', 'national_id')
        }),
        ('Additional Information', {
            'fields': ('english_name', 'photo', 'created')
        }),
    )

    def display_photo(self, obj):
        # عرض الصورة الشخصية في واجهة الإدارة
        if obj.photo and hasattr(obj.photo, 'url'):
            try:
                return format_html(
                    '<img src="{}" width="50" height="50" style="border-radius: 50%;" alt="صورة الطالب" />',
                    obj.photo.url
                )
            except Exception as e:
                return f"خطأ: {str(e)}"
        return "-"
    display_photo.short_description = "الصورة الشخصية"

    def has_english_name(self, obj):
        # التحقق من وجود اسم إنجليزي
        return bool(obj.english_name)
    has_english_name.boolean = True
    has_english_name.short_description = "يوجد اسم إنجليزي؟"

    def save_model(self, request, obj, form, change):
        # التحقق من الصورة عند حفظ النموذج
        if 'photo' in form.changed_data and obj.photo:
            # التحقق من حجم الصورة
            if obj.photo.size > 2 * 1024 * 1024:
                from django.contrib import messages
                messages.error(request, "الصورة أكبر من 2 ميجابايت، رجاءً اختر صورة أصغر.")
                return
            # التحقق من نوع الملف
            ext = obj.photo.name.split('.')[-1].lower()
            if ext not in ['jpg', 'jpeg', 'png']:
                from django.contrib import messages
                messages.error(request, "يُسمح فقط بصور بصيغة PNG أو JPG.")
                return
        super().save_model(request, obj, form, change)

    # def export_to_excel(self, request, queryset):
    #     # إنشاء ملف Excel جديد
    #     workbook = openpyxl.Workbook()
    #     worksheet = workbook.active
    #     worksheet.title = "Students"

    #     # تحديد رؤوس الأعمدة
    #     headers = [
    #         "رقم الجلوس",
    #         "كود الطالب",
    #         "الاسم العربي",
    #         "الرقم القومي",
    #         "الاسم الإنجليزي",
    #         "الصورة الشخصية",
    #         "يوجد اسم إنجليزي؟",
    #         "تاريخ الإنشاء"
    #     ]

    #     # كتابة الرؤوس في الصف الأول
    #     for col_num, header in enumerate(headers, 1):
    #         worksheet.cell(row=1, column=col_num).value = header
    #         worksheet.column_dimensions[get_column_letter(col_num)].width = 20

    #     # ضبط ارتفاع الصف الأول (الرأس)
    #     worksheet.row_dimensions[1].height = 20

    #     # ملء البيانات من السجلات المحددة
    #     for row_num, student in enumerate(queryset, 2):
    #         # كتابة البيانات النصية
    #         worksheet.cell(row=row_num, column=1).value = student.seat_number
    #         worksheet.cell(row=row_num, column=2).value = student.code
    #         worksheet.cell(row=row_num, column=3).value = student.arabic_name
    #         worksheet.cell(row=row_num, column=4).value = student.national_id
    #         worksheet.cell(row=row_num, column=5).value = student.english_name or "-"
    #         worksheet.cell(row=row_num, column=7).value = "نعم" if student.english_name else "لا"
    #         worksheet.cell(row=row_num, column=8).value = student.created.strftime('%Y-%m-%d %H:%M:%S')

    #         # إضافة الصورة الشخصية
    #         if student.photo and hasattr(student.photo, 'url'):
    #             try:
    #                 # الحصول على المسار الفعلي للصورة على الخادم
    #                 image_path = os.path.join(settings.MEDIA_ROOT, str(student.photo))
    #                 if os.path.exists(image_path):
    #                     # إنشاء كائن صورة
    #                     img = Image(image_path)
    #                     # ضبط حجم الصورة
    #                     img.width = 50
    #                     img.height = 50
    #                     # إضافة الصورة إلى الخلية
    #                     worksheet.add_image(img, f'{get_column_letter(6)}{row_num}')
    #                     # ضبط ارتفاع الصف لاستيعاب الصورة
    #                     worksheet.row_dimensions[row_num].height = 40
    #                 else:
    #                     worksheet.cell(row=row_num, column=6).value = "الصورة غير موجودة"
    #             except Exception as e:
    #                 worksheet.cell(row=row_num, column=6).value = f"خطأ: {str(e)}"
    #         else:
    #             worksheet.cell(row=row_num, column=6).value = "-"

    #     # إعداد الاستجابة لتحميل الملف
    #     response = HttpResponse(
    #         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #     )
    #     response['Content-Disposition'] = 'attachment; filename=students.xlsx'

    #     # حفظ الملف وإرساله في الاستجابة
    #     workbook.save(response)
    #     return response

    # export_to_excel.short_description = "تصدير إلى Excel"