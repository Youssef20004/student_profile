from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import ProfileForm
from .models import Student
import openpyxl
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import StudentSerializer
from urllib.parse import urljoin
from django.conf import settings

def login_view(request):
    if request.method == 'POST':
        seat_number = request.POST.get('seat_number')
        try:
            student = Student.objects.get(seat_number=seat_number)
            request.session['student_id'] = student.id
            return redirect('profile')
        except Student.DoesNotExist:
            messages.error(request, 'رقم الجلوس غير صحيح.')
    return render(request, 'index.html')

def profile_view(request):
    student_id = request.session.get('student_id')
    if not student_id:
        return redirect('login')
    
    student = Student.objects.get(id=student_id)
    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث البروفايل بنجاح!')
            return redirect('profile')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
    else:
        form = ProfileForm(instance=student)
    
    return render(request, 'profile.html', {'student': student, 'form': form})

class StudentLoginView(APIView):
    def post(self, request):
        seat_number = request.data.get('seat_number')
        if not seat_number:
            return Response(
                {'error': 'رقم الجلوس مطلوب.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            student = Student.objects.get(seat_number=seat_number)
            serializer = StudentSerializer(student, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Student.DoesNotExist:
            return Response(
                {'error': 'رقم الجلوس غير صحيح.'},
                status=status.HTTP_404_NOT_FOUND
            )

@staff_member_required
def export_students_to_excel(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Students"
    headers = ["رقم الجلوس", "كود الطالب", "الاسم العربي", "الرقم القومي", "الاسم الإنجليزي", "الصورة الشخصية", "يوجد اسم إنجليزي؟", "تاريخ الإنشاء"]
    worksheet.append(headers)
    students = Student.objects.all()
    base_url = request.build_absolute_uri('/')
    for row_num, student in enumerate(students, 2):
        photo_url = urljoin(base_url, student.photo.url) if student.photo and hasattr(student.photo, 'url') else "-"
        worksheet.cell(row=row_num, column=1).value = student.seat_number
        worksheet.cell(row=row_num, column=2).value = student.code
        worksheet.cell(row=row_num, column=3).value = student.arabic_name
        worksheet.cell(row=row_num, column=4).value = student.national_id
        worksheet.cell(row=row_num, column=5).value = student.english_name or "-"
        worksheet.cell(row=row_num, column=6).value = photo_url
        worksheet.cell(row=row_num, column=7).value = "نعم" if student.english_name else "لا"
        worksheet.cell(row=row_num, column=8).value = student.created.strftime('%Y-%m-%d %H:%M:%S')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    workbook.save(response)
    return response

class StudentDetailView(APIView):
    def get_object(self, identifier):
        try:
            if identifier.isdigit() and len(identifier) == 14:
                return Student.objects.get(national_id=identifier)
            return Student.objects.get(seat_number=identifier)
        except Student.DoesNotExist:
            return None

    def get(self, request, identifier):
        student = self.get_object(identifier)
        if not student:
            return Response(
                {'error': 'الطالب غير موجود.'},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = StudentSerializer(student, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, identifier):
        student = self.get_object(identifier)
        if not student:
            return Response(
                {'error': 'الطالب غير موجود.'},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = StudentSerializer(
            student,
            data=request.data,
            partial=True,
            context={'request': request}
        )
        if serializer.is_valid():
            serializer.save()
            return Response(
                {'message': 'تم تحديث البروفايل بنجاح!', 'data': serializer.data},
                status=status.HTTP_200_OK
            )
        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )
