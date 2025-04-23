import openpyxl
import random

# قوايم لتوليد أسماء عربية
first_names = ['أحمد', 'محمد', 'سارة', 'فاطمة', 'علي', 'نورا', 'خالد', 'منى', 'يوسف', 'ليلى', 'حسن', 'زينب', 'عمر', 'ريم']
middle_names = ['محمد', 'عبد الله', 'محمود', 'خالد', 'حسين', 'أحمد', 'إبراهيم', 'علي', 'سمير']
last_names = ['علي', 'حسن', 'أحمد', 'عبد الرحمن', 'سعيد', 'صالح', 'عمر', 'محمود', 'حسن']

# إنشاء ملف Excel
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Students"

# رؤوس الأعمدة
headers = ['seat_number', 'code', 'arabic_name', 'national_id']
for col_num, header in enumerate(headers, 1):
    worksheet.cell(row=1, column=col_num).value = header
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

# توليد 1000 سجل
for i in range(100000):
    seat_number = 100001 + i
    code = f'ST{str(i+1).zfill(3)}'  # مثل ST001, ST002
    arabic_name = f"{random.choice(first_names)} {random.choice(middle_names)} {random.choice(last_names)}"
    # توليد رقم قومي وهمي (14 رقم)
    national_id = ''.join([str(random.randint(0, 9)) for _ in range(14)])
    
    # كتابة البيانات في الصف
    row = [seat_number, code, arabic_name, national_id]
    for col_num, value in enumerate(row, 1):
        worksheet.cell(row=i+2, column=col_num).value = value

# حفظ الملف
workbook.save('students_1000.xlsx')
print("تم إنشاء ملف students_1000.xlsx بنجاح!")